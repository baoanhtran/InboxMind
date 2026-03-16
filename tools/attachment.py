"""
tools/attachment.py

Sends attachments directly to GPT-4o in binary form where the OpenAI API
supports it, falling back to local text extraction for formats it doesn't.

OpenAI native binary support (as of 2025):
  - Images (JPEG, PNG, GIF, WEBP) → base64 data URI, type "image_url"
  - PDF                            → base64, type "file" with source_type "base64"

No native binary support (local extraction used instead):
  - Excel / XLSX   → openpyxl → text
  - CSV            → decoded as text
  - Plain text     → decoded as text
"""

from __future__ import annotations

import base64
import csv
import io
from typing import Optional

from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage


_ANALYSIS_PROMPT = (
    "Analyse this file thoroughly. "
    "Extract all text, tables, figures, and data you can see. "
    "Identify the key facts, numbers, names, dates, and decisions. "
    "Focus on information that would be relevant in a business email context."
)


# ---------------------------------------------------------------------------
# Dispatcher
# ---------------------------------------------------------------------------

def extract_attachment_text(
    raw_bytes: bytes,
    filename: str,
    mime_type: str,
    llm: Optional[ChatOpenAI] = None,
) -> str:
    """
    Process an email attachment and return its content as text.

    For images and PDFs, the raw bytes are sent directly to GPT-4o
    without any intermediate extraction step.

    For Excel, CSV, and plain text, local libraries handle extraction
    because the OpenAI API has no native understanding of those formats.
    """
    mime_lower = mime_type.lower()
    name_lower = filename.lower()

    if _is_image(mime_lower, name_lower):
        if llm is not None:
            print(f"[attachment] Sending image '{filename}' directly to GPT-4o Vision")
            return _send_image_to_gpt4o(raw_bytes, mime_type, llm)
        return "[Image attachment — no LLM provided for Vision analysis]"

    if "pdf" in mime_lower or name_lower.endswith(".pdf"):
        if llm is not None:
            print(f"[attachment] Sending PDF '{filename}' directly to GPT-4o")
            return _send_pdf_to_gpt4o(raw_bytes, llm)
        print(f"[attachment] No LLM provided, falling back to local PDF extraction for '{filename}'")
        return _extract_pdf_locally(raw_bytes)

    if _is_excel(mime_lower, name_lower):
        print(f"[attachment] Extracting Excel '{filename}' locally (no OpenAI native support)")
        return _extract_excel(raw_bytes)

    if name_lower.endswith(".csv"):
        print(f"[attachment] Extracting CSV '{filename}' locally")
        return _extract_csv(raw_bytes)

    if "text" in mime_lower or name_lower.endswith(".txt"):
        return raw_bytes.decode("utf-8", errors="replace")

    return f"[Unsupported attachment type: {mime_type} / {filename}]"


def _is_image(mime_lower: str, name_lower: str) -> bool:
    image_mimes = ("image/jpeg", "image/png", "image/gif", "image/webp")
    image_exts  = (".jpg", ".jpeg", ".png", ".gif", ".webp")
    return any(m in mime_lower for m in image_mimes) or name_lower.endswith(image_exts)


def _is_excel(mime_lower: str, name_lower: str) -> bool:
    excel_mimes = ("spreadsheet", "excel", "xlsx", "xls")
    excel_exts  = (".xlsx", ".xls")
    return any(x in mime_lower for x in excel_mimes) or name_lower.endswith(excel_exts)


# ---------------------------------------------------------------------------
# Images → GPT-4o Vision (binary via base64 data URI)
# ---------------------------------------------------------------------------

def _send_image_to_gpt4o(raw_bytes: bytes, mime_type: str, llm: ChatOpenAI) -> str:
    """
    Send image bytes directly to GPT-4o Vision using a base64 data URI.
    No pre-processing — the model receives the raw image.
    """
    try:
        b64 = base64.standard_b64encode(raw_bytes).decode("utf-8")

        message = HumanMessage(
            content=[
                {"type": "text", "text": _ANALYSIS_PROMPT},
                {
                    "type": "image_url",
                    "image_url": {
                        "url":    f"data:{mime_type};base64,{b64}",
                        "detail": "high",
                    },
                },
            ]
        )

        response = llm.invoke([message])
        return response.content

    except Exception as exc:
        return f"[GPT-4o Vision error: {exc}]"


# ---------------------------------------------------------------------------
# PDF → GPT-4o (binary via base64 file content block)
# ---------------------------------------------------------------------------

def _send_pdf_to_gpt4o(raw_bytes: bytes, llm: ChatOpenAI) -> str:
    """
    Send a PDF to GPT-4o via the OpenAI Files API.

    The correct flow for PDFs is:
      1. Upload the file via POST /v1/files (purpose="assistants") to get a file_id
      2. Reference that file_id in the message content block
      3. Delete the file afterwards to avoid storage accumulation

    Falls back to local PyMuPDF extraction if the upload fails.
    """
    import os
    import httpx

    api_key = os.environ["OPENAI_API_KEY"]

    try:
        # Step 1: upload the file to get a file_id
        upload_response = httpx.post(
            "https://api.openai.com/v1/files",
            headers={"Authorization": f"Bearer {api_key}"},
            files={"file": ("attachment.pdf", raw_bytes, "application/pdf")},
            data={"purpose": "assistants"},
            timeout=60,
        )
        upload_response.raise_for_status()
        file_id = upload_response.json()["id"]

        # Step 2: send the file_id to GPT-4o
        message = HumanMessage(
            content=[
                {"type": "text", "text": _ANALYSIS_PROMPT},
                {
                    "type": "file",
                    "file": {"file_id": file_id},
                },
            ]
        )
        response = llm.invoke([message])
        result = response.content

        # Step 3: delete the uploaded file to avoid storage buildup
        try:
            httpx.delete(
                f"https://api.openai.com/v1/files/{file_id}",
                headers={"Authorization": f"Bearer {api_key}"},
                timeout=10,
            )
        except Exception:
            pass  # non-fatal if delete fails

        return result

    except Exception as exc:
        print(f"[attachment] GPT-4o PDF API error ({exc}), falling back to local extraction")
        return _extract_pdf_locally(raw_bytes)


# ---------------------------------------------------------------------------
# PDF local fallback (PyMuPDF)
# ---------------------------------------------------------------------------

def _extract_pdf_locally(raw_bytes: bytes) -> str:
    """Fallback: extract text from PDF using PyMuPDF."""
    try:
        import fitz

        doc = fitz.open(stream=raw_bytes, filetype="pdf")
        pages: list[str] = []
        for page_num in range(min(doc.page_count, 30)):
            page = doc[page_num]
            pages.append(f"--- Page {page_num + 1} ---\n{page.get_text()}")
        doc.close()
        return "\n".join(pages)

    except ImportError:
        return "[PyMuPDF not installed and GPT-4o PDF failed — cannot extract PDF]"
    except Exception as exc:
        return f"[Local PDF extraction error: {exc}]"


# ---------------------------------------------------------------------------
# Excel (local only — no OpenAI native binary support)
# ---------------------------------------------------------------------------

def _extract_excel(raw_bytes: bytes) -> str:
    """Extract content from .xlsx using openpyxl."""
    try:
        import openpyxl

        wb     = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        output: list[str] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output.append(f"=== Sheet: {sheet_name} ===")
            for row in ws.iter_rows(max_row=200, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                output.append("\t".join(cells))

        return "\n".join(output)

    except ImportError:
        return "[openpyxl not installed — cannot extract Excel. Run: pip install openpyxl]"
    except Exception as exc:
        return f"[Excel extraction error: {exc}]"


# ---------------------------------------------------------------------------
# CSV (local only)
# ---------------------------------------------------------------------------

def _extract_csv(raw_bytes: bytes) -> str:
    """Decode and return CSV content as plain text."""
    try:
        text   = raw_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows   = ["\t".join(row) for row in reader]
        return "\n".join(rows[:500])
    except Exception as exc:
        return f"[CSV extraction error: {exc}]"